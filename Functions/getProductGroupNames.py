def getProductGroupNames(base):
    TYPES = 'MWC'
    # Функция составляет словарь соответствия наименования товарных групп 1 или 2 в 1С, таблицах и на сайте
    if base not in TYPES:
        raise ValueError('Базовый показатель для определния словаря соответствий указан неверно! Введите "C", "M" или "W"!')
    import openpyxl
    from openpyxl.utils import get_column_letter as colLetter
    from openpyxl.utils import coordinate_from_string, column_index_from_string
    from Config.constants import DATATABLE
    TYPES = TYPES.replace(base, '')
    workBook = openpyxl.load_workbook(DATATABLE)
    workPage = workBook.active
    HEADERS = {}
    RESULT = {}
    for cell in workPage[1]:
        HEADERS[cell.value] = column_index_from_string(cell.column)
    for i in range(2, workPage.max_row+1):
        currentValue = workPage.cell(row = i, column = HEADERS[base]).value
        RESULT[currentValue] = {}
        RESULT[currentValue]['ID'] = workPage.cell(row = i, column = HEADERS['ID']).value
        RESULT[currentValue]['N'] = workPage.cell(row = i, column = HEADERS['N']).value
        RESULT[currentValue]['W'] = workPage.cell(row = i, column = HEADERS['W']).value
        RESULT[currentValue]['M'] = workPage.cell(row = i, column = HEADERS['M']).value
        RESULT[currentValue]['C'] = workPage.cell(row = i, column = HEADERS['C']).value
    return RESULT